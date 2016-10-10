#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import Workbook

class exelbook():
	def __init__(self):
		self.Workbook = Workbook()
		self.Worksheet = self.Workbook.active
		
	def setTitle(self, title):
		self.Worksheet.title = title
		
	def setColum(self, columnstr, data):
		self.Worksheet[columnstr] = data
		
	def setRow(self, number, datalist):
		for i in range(0, len(datalist)):
			key = chr(i + ord("A")) + str(number)
			#print(key)
			self.Worksheet[key] = datalist[i]
		
	def setLine(self, letter, datalist): #den här sätter alltså upifrån och ner
		for i in range(0, len(datalist)):
			key = letter + str(i+1) #för att excels celler börjar på 1 och inte 0
			#print(key)
			self.Worksheet[key] = datalist[i]
			
	def saveas(self, filename):
		self.filename = filename
		self.Workbook.save(self.filename)
		
	def save(self):
		self.filename = self.Worksheet.title + ".xlsx"
		self.Workbook.save(self.filename)
		
	def advancedwrite(self, listoflists):
		currentRow = 1
		self.setRow(currentRow, ["In/Ut", "Datum", "Stampel", "Lektionstid", "Franvaro", "Anm"])
		currentRow = currentRow + 1
		for i in listoflists:
			self.setRow(currentRow, i)
			currentRow = currentRow + 1


		
#muhbook = exelbook()
#muhbook.setTitle("Jakob Christoffersson")
#muhbook.advancedwrite([["1","2","3"],["3","2","1"]])

#ws['A1'] = "A1"
#ws['A2'] = "A2"

#muhbook.save()